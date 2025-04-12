from aiogram import Bot, Dispatcher, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
from docx import Document
import asyncio
import os
import json
import logging
import time
import random
from datetime import datetime

from quiz_utils import convert_format, calculate_points, get_result_message, parse_text_file
from storage import TestStorage
from localization import get_text
from database import init_db

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Bot configuration
TOKEN = "8184215515:AAEVINsnkj_fTBbxZfBpvqZtUCsNj2kvwjo"  # @BotFather dan olgan tokeningizni kiriting
ADMIN_CHANNEL = "englishpodcasts_panorama"  # Admin channel ID ni kiriting
FEEDBACK_CHANNEL = "usercommentss"  # Shikoyatlar kanali ID
BOT_USERNAME = "masterquizz_bot"  # Bot username (https://t.me/masterquizz_bot)
ADMIN_IDS = [1477944238]  # Admin ID larini kiriting
# Add flag to track if user has seen the referral link
USER_FIRST_JOIN = {}  # Store user_id -> True/False
bot = Bot(token=TOKEN)
dp = Dispatcher()

# Initialize test storage
test_storage = TestStorage()

class UserData:
    def __init__(self):
        self.users = {}  # {user_id: {"username": str, "full_name": str, "joined_date": str, "language": str}}
        self.total_quizzes = 0

user_data = UserData()

class QuizStates(StatesGroup):
    waiting_for_language = State()
    waiting_for_file = State()
    waiting_for_file_name = State()
    waiting_for_range = State()
    waiting_for_shuffle = State()
    waiting_for_quiz = State()
    in_quiz = State()
    selecting_test = State()
    waiting_for_feedback = State()
    invite_friends = State()
    # New states for broadcasting
    broadcast_selecting_type = State()
    broadcast_waiting_text = State()
    broadcast_waiting_photo = State()
    broadcast_waiting_video = State()
    broadcast_waiting_poll_question = State()
    broadcast_waiting_poll_options = State()
    broadcast_confirming = State()

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

class UserScore:
    def __init__(self):
        self.scores = {}  # {user_id: {correct: X, total: Y}}

    def update_score(self, user_id: int, is_correct: bool):
        if user_id not in self.scores:
            self.scores[user_id] = {"correct": 0, "total": 0}
        
        self.scores[user_id]["total"] += 1
        if is_correct:
            self.scores[user_id]["correct"] += 1

    def get_score(self, user_id: int):
        if user_id not in self.scores:
            return 0, 0
        return self.scores[user_id]["correct"], self.scores[user_id]["total"]

user_scores = UserScore()

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    from database import add_user, add_referral, has_invited_friend, get_referrer
    
    user_id = message.from_user.id
    is_new_user = user_id not in user_data.users
    
    # Parse start command arguments for referral
    referrer_id = None
    start_args = message.text.split()
    if len(start_args) > 1:
        # Format is /start ref123456789
        try:
            referral_code = start_args[1]
            if referral_code.startswith('ref'):
                referrer_id = int(referral_code[3:])
                logger.info(f"User {user_id} was referred by {referrer_id}")
        except Exception as e:
            logger.error(f"Error parsing referral code: {e}")
    
    # Add user to memory
    if is_new_user:
        user_data.users[user_id] = {
            "username": message.from_user.username,
            "full_name": message.from_user.full_name,
            "joined_date": message.date.strftime("%Y-%m-%d %H:%M:%S"),
            "language": "uz"  # Default til
        }
    
    # Add user to database
    add_user(
        user_id=user_id,
        username=message.from_user.username,
        first_name=message.from_user.first_name,
        last_name=message.from_user.last_name
    )
    
    # Process referral if new user and referrer_id is valid
    if is_new_user and referrer_id and referrer_id != user_id:
        # Add referral record
        success = add_referral(referrer_id, user_id)
        if success:
            # Notify referrer
            try:
                referrer_lang = await get_user_language(referrer_id)
                await bot.send_message(
                    referrer_id, 
                    get_text(referrer_lang, "referral_success"),
                    parse_mode="HTML"
                )
            except Exception as e:
                logger.error(f"Failed to notify referrer: {e}")
    
    # Only show referral link if user is new (first time joining)
    if is_new_user:
        user_referral_link = f"https://t.me/{BOT_USERNAME}?start=ref{user_id}"
        USER_FIRST_JOIN[user_id] = True
        
        user_lang = await get_user_language(user_id)
        await message.answer(
            get_text(user_lang, "your_invite_link").format(link=user_referral_link),
            parse_mode="HTML"
        )
    
    # Tilni tanlash
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(text="🇺🇿 O'zbek tili", callback_data="language:uz"),
                types.InlineKeyboardButton(text="🇷🇺 Русский язык", callback_data="language:ru")
            ]
        ]
    )
    
    await message.answer(get_text("uz", "select_language"), reply_markup=keyboard, parse_mode="HTML")
    await state.set_state(QuizStates.waiting_for_language)

@dp.callback_query(lambda c: c.data.startswith("language:"))
async def language_selected(callback_query: types.CallbackQuery, state: FSMContext):
    from database import has_invited_friend
    
    user_id = callback_query.from_user.id
    lang = callback_query.data.split(':')[1]
    
    # Update user language
    if user_id in user_data.users:
        user_data.users[user_id]["language"] = lang
    
    await callback_query.answer()
    await callback_query.message.delete()
    
    # Send confirmation message
    await callback_query.message.answer(get_text(lang, "language_selected"))
    
    # Check if user has already invited someone or is an admin
    if is_admin(user_id) or has_invited_friend(user_id):
        # User can access the bot normally
        if has_invited_friend(user_id) and not is_admin(user_id):
            await callback_query.message.answer(get_text(lang, "already_invited"))
        
        # Show main menu with selected language
        await show_main_menu(callback_query.message, lang)
        await state.clear()
    else:
        # User needs to invite a friend but we don't send the link again
        # since it was already sent in the /start command
        await callback_query.message.answer(get_text(lang, "need_invite_friend"))
        await state.clear()

async def get_user_language(user_id):
    """Get user language or default to Uzbek"""
    if user_id in user_data.users and "language" in user_data.users[user_id]:
        return user_data.users[user_id]["language"]
    return "uz"  # Default is Uzbek

async def show_main_menu(message: types.Message, lang=None):
    """Show main menu with language-specific buttons"""
    if lang is None:
        lang = await get_user_language(message.from_user.id)
    
    keyboard_buttons = [
        [
            types.KeyboardButton(text=get_text(lang, "btn_create_quiz")), 
            types.KeyboardButton(text=get_text(lang, "btn_my_tests"))
        ],
        [
            types.KeyboardButton(text=get_text(lang, "btn_results")), 
            types.KeyboardButton(text=get_text(lang, "btn_guide"))
        ],
        [
            types.KeyboardButton(text=get_text(lang, "btn_feedback")),
            types.KeyboardButton(text=get_text(lang, "btn_invite"))
        ]
    ]
    
    if is_admin(message.from_user.id):
        keyboard_buttons.append([
            types.KeyboardButton(text=get_text(lang, "btn_admin_stats")),
            types.KeyboardButton(text=get_text(lang, "btn_broadcast"))
        ])
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=keyboard_buttons,
        resize_keyboard=True,
        input_field_placeholder=get_text(lang, "menu_placeholder")
    )
    
    await message.answer(get_text(lang, "bot_welcome"), reply_markup=keyboard, parse_mode="HTML")

# Button handlers for create quiz
@dp.message(lambda message: message.text == get_text("uz", "btn_create_quiz") or 
                         message.text == get_text("ru", "btn_create_quiz"))
async def quiz_create(message: types.Message, state: FSMContext):
    from database import has_invited_friend
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Check if user has invited friends or is admin
    if is_admin(user_id) or has_invited_friend(user_id):
        # Updated message to mention both .docx and .txt support
        await message.answer(get_text(lang, "upload_file"))
        await state.set_state(QuizStates.waiting_for_file)
    else:
        # User needs to invite a friend first
        await message.answer(get_text(lang, "need_invite_friend"))
        await invite_friends(message)

# Button handlers for my results
@dp.message(lambda message: message.text == get_text("uz", "btn_results") or 
                         message.text == get_text("ru", "btn_results"))
async def show_results(message: types.Message):
    from database import has_invited_friend
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Check if user has invited friends or is admin
    if not (is_admin(user_id) or has_invited_friend(user_id)):
        # User needs to invite a friend first
        await message.answer(get_text(lang, "need_invite_friend"))
        await invite_friends(message)
        return
    
    # Check if user has any test results
    if user_id not in user_data.users or "test_results" not in user_data.users[user_id] or not user_data.users[user_id]["test_results"]:
        await message.answer(get_text(lang, "no_results"))
        return
    
    # Get user's test results (simplified list)
    results = user_data.users[user_id]["test_results"]
    results_list = []
    
    for test_result in results:
        test_name = test_result.get("test_name", "Test")
        date = test_result.get("date", "")
        correct = test_result.get("correct", 0)
        total = test_result.get("total", 0)
        percent = round((correct / total * 100), 1) if total > 0 else 0
        
        # Create simplified result item using the template from localization
        result_item = get_text(lang, "quiz_results_list_item").format(
            name=test_name,
            date=date,
            percent=percent,
            correct=correct,
            total=total
        )
        results_list.append(result_item)
    
    # Combine all results
    combined_results = "\n\n".join(results_list)
    final_message = f"<b>📊 {get_text(lang, 'btn_results')}</b>\n\n{combined_results}"
    
    await message.answer(final_message, parse_mode="HTML")

# Button handlers for admin statistics
@dp.message(lambda message: message.text == get_text("uz", "btn_admin_stats") or 
                         message.text == get_text("ru", "btn_admin_stats"))
async def admin_statistics(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    
    lang = await get_user_language(message.from_user.id)
    total_users = len(user_data.users)
    
    # Count total quizzes across all users
    total_quizzes = 0
    for user_id_str in test_storage.tests:
        total_quizzes += len(test_storage.tests[user_id_str])
    
    # Update internal counter
    user_data.total_quizzes = total_quizzes
    
    # Create stats summary text
    stats_text = get_text(lang, "stats_title") + "\n\n"
    stats_text += get_text(lang, "stats_general").format(users_count=total_users, tests_count=total_quizzes)
    
    # Send stats summary first
    await message.answer(stats_text, parse_mode="HTML")
    
    # Create users list as file content
    users_text = get_text(lang, "stats_users_title") + "\n\n"
    
    for user_id, data in user_data.users.items():
        lang_info = f" - {data.get('language', 'uz')}" if 'language' in data else ""
        users_text += f"- {data['full_name']} (@{data['username']}){lang_info}\n" \
                      f"  ID: {user_id}\n" \
                      f"  Joined: {data['joined_date']}\n\n"
    
    # Create and send users list as a text file
    import io
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Create Excel file for better formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Add header with formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Set header cells
    headers = ["Name", "Username", "ID", "Language", "Joined Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add user data
    row = 2
    for user_id, data in user_data.users.items():
        ws.cell(row=row, column=1, value=data.get('full_name', 'Unknown'))
        ws.cell(row=row, column=2, value=f"@{data.get('username', 'noname')}")
        ws.cell(row=row, column=3, value=str(user_id))
        ws.cell(row=row, column=4, value=data.get('language', 'uz'))
        ws.cell(row=row, column=5, value=data.get('joined_date', 'Unknown'))
        
        # Apply border to cells
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save Excel file to BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Also create text file as alternative
    file_name_txt = f"users_list_{today}.txt"
    users_file_txt = io.BytesIO(users_text.encode('utf-8'))
    
    # Send both files - Excel and text
    await message.answer_document(
        types.BufferedInputFile(
            file=excel_file.getvalue(),
            filename=f"users_list_{today}.xlsx"
        ),
        caption=f"📊 Excel {get_text(lang, 'user_count').format(count=total_users)}"
    )
    
    await message.answer_document(
        types.BufferedInputFile(
            file=users_file_txt.getvalue(),
            filename=file_name_txt
        ),
        caption=f"📝 Text {get_text(lang, 'user_count').format(count=total_users)}"
    )

# Button handlers for my tests
@dp.message(lambda message: message.text == get_text("uz", "btn_my_tests") or 
                         message.text == get_text("ru", "btn_my_tests"))
async def show_my_tests(message: types.Message, state: FSMContext):
    from database import has_invited_friend
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Check if user has invited friends or is admin
    if is_admin(user_id) or has_invited_friend(user_id):
        tests = test_storage.get_user_tests(user_id)
        
        if not tests:
            await message.answer(get_text(lang, "no_tests"))
            return
    else:
        # User needs to invite a friend first
        await message.answer(get_text(lang, "need_invite_friend"))
        await invite_friends(message)
        return
    
    # Create inline keyboard with test names
    buttons = []
    for i, test in enumerate(tests):
        buttons.append([types.InlineKeyboardButton(
            text=f"{i+1}. {test['name']} ({len(test['questions'])} {'savol' if lang == 'uz' else 'вопр.'})",
            callback_data=f"select_test:{i}"
        )])
    
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await message.answer(get_text(lang, "available_tests"), reply_markup=keyboard, parse_mode="HTML")
    await state.set_state(QuizStates.selecting_test)

@dp.callback_query(lambda c: c.data.startswith("select_test:"))
async def process_test_selection(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    test = test_storage.get_test(user_id, test_index)
    
    if not test:
        await callback_query.message.answer(get_text(lang, "test_not_found"))
        return
    
    # Show test details and options with improved styling
    buttons = [
        [
            types.InlineKeyboardButton(text=get_text(lang, "btn_start_test"), callback_data=f"start_test:{test_index}"),
            types.InlineKeyboardButton(text=get_text(lang, "btn_delete_test"), callback_data=f"delete_test:{test_index}")
        ],
        [types.InlineKeyboardButton(text=get_text(lang, "btn_back"), callback_data="back_to_tests")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    
    test_info = get_text(lang, "test_info").format(
        name=test['name'], 
        question_count=len(test['questions']), 
        created_at=test['created_at']
    )
    
    await callback_query.message.answer(test_info, reply_markup=keyboard, parse_mode="HTML")

@dp.callback_query(lambda c: c.data.startswith("start_test:"))
async def start_saved_test(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    test = test_storage.get_test(user_id, test_index)
    
    if not test:
        await callback_query.message.answer(get_text(lang, "test_not_found"))
        return
    
    # Ask for range
    questions = test["questions"]
    await state.update_data(
        questions=questions,
        test_name=test["name"]
    )
    
    await callback_query.message.answer(
        f"📚 {test['name']}: {len(questions)} {'savol' if lang == 'uz' else 'вопросов'}.\n"
        f"{get_text(lang, 'test_saved').format(name=test['name'], count=len(questions))}"
    )
    await state.set_state(QuizStates.waiting_for_range)

@dp.callback_query(lambda c: c.data.startswith("delete_test:"))
async def delete_saved_test(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    success = test_storage.delete_test(user_id, test_index)
    
    if success:
        await callback_query.message.answer(get_text(lang, "test_deleted"))
    else:
        await callback_query.message.answer(get_text(lang, "test_delete_error"))
    
    # Show updated tests list
    await show_my_tests(callback_query.message, state)

@dp.callback_query(lambda c: c.data == "back_to_tests")
async def back_to_tests_list(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    await show_my_tests(callback_query.message, state)

# Button handlers for guide (renamed from help)
@dp.message(lambda message: message.text == get_text("uz", "btn_guide") or 
                         message.text == get_text("ru", "btn_guide"))
async def show_guide(message: types.Message):
    # Guide is available to all users without any restrictions,
    # so no need to check for invites
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    help_title = get_text(lang, "help_title")
    help_text = get_text(lang, "help_text")
    
    # First, send the guide text
    await message.answer(f"{help_title}\n\n{help_text}", parse_mode="HTML")
    
    # Then, send the video guide if it exists
    manual_video_path = "media/manual.mp4"
    if os.path.exists(manual_video_path) and os.path.getsize(manual_video_path) > 0:
        try:
            with open(manual_video_path, 'rb') as video_file:
                await message.answer_video(
                    video=types.BufferedInputFile(
                        file=video_file.read(),
                        filename="manual.mp4"
                    ),
                    caption=get_text(lang, "video_guide_caption") if lang in ["uz", "ru"] else "Video guide on how to use the bot"
                )
            logger.info(f"Sent guide video to user {message.from_user.id}")
        except Exception as e:
            logger.error(f"Failed to send guide video: {e}")
            await message.answer(
                get_text(lang, "video_guide_error") if lang in ["uz", "ru"] else 
                "Sorry, there was an error sending the video guide. Please try again later."
            )

# Add command for user count
@dp.message(Command("users"))
async def cmd_users(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    
    lang = await get_user_language(message.from_user.id)
    user_count = len(user_data.users)
    await message.answer(get_text(lang, "user_count").format(count=user_count))

# Add feedback feature
@dp.message(lambda message: message.text == get_text("uz", "btn_feedback") or 
                         message.text == get_text("ru", "btn_feedback"))
async def request_feedback(message: types.Message, state: FSMContext):
    from database import has_invited_friend
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Check if user has invited friends or is admin
    if not (is_admin(user_id) or has_invited_friend(user_id)):
        # User needs to invite a friend first
        await message.answer(get_text(lang, "need_invite_friend"))
        await invite_friends(message)
        return
    
    await message.answer(get_text(lang, "feedback_prompt"))
    await state.set_state(QuizStates.waiting_for_feedback)

@dp.message(QuizStates.waiting_for_feedback)
async def handle_feedback(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    feedback_text = message.text
    
    # Send confirmation to user
    await message.answer(get_text(lang, "feedback_sent"))
    
    # User haqida ma'lumot
    user_info = user_data.users.get(user_id, {"full_name": "Foydalanuvchi", "username": ""})
    feedback_message = get_text(lang, "user_feedback").format(
        message=feedback_text,
        name=user_info["full_name"],
        username=user_info["username"] or "noname"
    )
    
    # Send feedback to dedicated feedback channel first (using username-based channel)
    if FEEDBACK_CHANNEL:
        try:
            await bot.send_message("@" + FEEDBACK_CHANNEL, feedback_message, parse_mode="HTML")
            logger.info(f"Feedback sent to @{FEEDBACK_CHANNEL} from user {user_id}")
        except Exception as e:
            logger.error(f"Failed to send feedback to feedback channel: {e}")
            # If error occurs, send to admin channel as backup
            if ADMIN_CHANNEL:
                try:
                    await bot.send_message("@" + ADMIN_CHANNEL, feedback_message, parse_mode="HTML")
                    logger.info(f"Feedback sent to admin channel (backup) from user {user_id}")
                except Exception as e2:
                    logger.error(f"Also failed to send feedback to admin channel: {e2}")
    
    # If no feedback channel, send to admin channel
    elif ADMIN_CHANNEL:
        try:
            await bot.send_message("@" + ADMIN_CHANNEL, feedback_message, parse_mode="HTML")
            logger.info(f"Feedback sent to admin channel from user {user_id}")
        except Exception as e:
            logger.error(f"Failed to send feedback to admin channel: {e}")
    
    await state.clear()
    await show_main_menu(message)

# Score command
@dp.message(Command("score"))
async def cmd_score(message: types.Message):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    correct, total = user_scores.get_score(user_id)
    
    if total == 0:
        await message.answer(get_text(lang, "no_results"))
    else:
        result_message = get_result_message(correct, total)
        await message.answer(result_message, parse_mode="HTML")

# Add command to stop quiz
@dp.message(Command("stop"))
async def cmd_stop_quiz(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    
    if current_state != QuizStates.in_quiz.state:
        return
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    
    current_question = data.get('current_question', 0)
    total_questions = data.get('total_questions', 0)
    correct_answers = data.get('correct_answers', 0)
    
    # Generate partial results
    result_message = get_text(lang, "test_stopped")
    if current_question > 0:
        result_message += get_result_message(correct_answers, current_question)
    
    # Add return to main menu button
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))]
        ],
        resize_keyboard=True
    )
    
    await message.answer(result_message, reply_markup=keyboard, parse_mode="HTML")
    await state.clear()

@dp.poll_answer()
async def handle_poll_answer(poll_answer: types.PollAnswer, state: FSMContext):
    user_id = poll_answer.user.id
    lang = await get_user_language(user_id)
    
    # Get current state data
    current_state = await state.get_state()
    if current_state != QuizStates.in_quiz.state:
        return
    
    data = await state.get_data()
    
    # Get the correct option ID from state (previously saved in send_quiz_question)
    correct_option_id = data.get('current_correct_option_id', 0)
    
    # Check if user selected the correct option
    # poll_answer.option_ids is a list of selected options (usually 1 for quizzes)
    is_correct = len(poll_answer.option_ids) > 0 and poll_answer.option_ids[0] == correct_option_id
    
    # Log the answer and correctness
    logger.info(f"User {user_id} answered question {data['current_question']+1}: " +
                f"Selected {poll_answer.option_ids[0]}, Correct: {correct_option_id}, " +
                f"Result: {'✓' if is_correct else '✗'}")
    
    # Update user score
    user_scores.update_score(user_id, is_correct)
    
    # Update question counters
    current_question = data['current_question'] + 1
    total_questions = data['total_questions']
    correct_answers = data['correct_answers'] + (1 if is_correct else 0)
    test_name = data.get('test_name', 'Test')
    
    # Save updated data
    await state.update_data(
        current_question=current_question,
        correct_answers=correct_answers
    )
    
    if current_question < total_questions:
        # Send next question
        await send_quiz_question(user_id, state)
    else:
        # Quiz finished, generate detailed results
        wrong_answers = total_questions - correct_answers
        percentage = round((correct_answers / total_questions * 100), 1) if total_questions > 0 else 0
        points_100 = calculate_points(correct_answers, total_questions, 100)
        
        # Current date and time for result
        from datetime import datetime
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Generate detailed result message
        detailed_result = get_text(lang, "quiz_detailed_results").format(
            name=test_name,
            date=current_date,
            correct=correct_answers,
            wrong=wrong_answers,
            total=total_questions,
            percent=percentage,
            points=points_100
        )
        
        # Add return to main menu button with improved styling
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))]
            ],
            resize_keyboard=True,
            input_field_placeholder=get_text(lang, "quiz_finish_placeholder")
        )
        
        # Send detailed results to user
        await bot.send_message(
            user_id,
            detailed_result,
            reply_markup=keyboard,
            parse_mode="HTML"
        )
        
        # Save result to user data for history
        if user_id not in user_data.users:
            user_data.users[user_id] = {}
        
        if "test_results" not in user_data.users[user_id]:
            user_data.users[user_id]["test_results"] = []
            
        # Add this test result to history
        user_data.users[user_id]["test_results"].append({
            "test_name": test_name,
            "date": current_date,
            "correct": correct_answers,
            "total": total_questions,
            "percent": percentage,
            "points": points_100
        })
        
        await state.clear()
    
    # Document is forwarded automatically by Telegram, 
    # but we no longer need to forward answers to admin channel

# Return to main menu button handler
@dp.message(lambda message: message.text == get_text("uz", "btn_main_menu") or 
                         message.text == get_text("ru", "btn_main_menu"))
async def return_to_main_menu(message: types.Message):
    lang = await get_user_language(message.from_user.id)
    await show_main_menu(message, lang)

async def send_quiz_question(user_id, state):
    """Send a quiz question to the user"""
    lang = await get_user_language(user_id)
    data = await state.get_data()
    current_question = data['current_question']
    total_questions = data['total_questions']
    
    # Always use quiz_questions which contains the final version of questions (shuffled or not)
    questions = data.get('quiz_questions', [])
    if not questions:
        logger.error(f"No quiz_questions found for user {user_id}")
        return
    
    shuffle_answers = data.get('shuffle_answers', False)
    
    # Get the current question and its options
    try:
        question, options = questions[current_question]
        # Always preserve the correct answer which is at index 0
        correct_answer = options[0]
        all_options = options.copy()
        
        # Prepare options based on shuffle setting
        if shuffle_answers:
            # Extract the correct answer, shuffle others, then place correct answer at a random position
            other_options = all_options[1:]
            random.shuffle(other_options)
            
            # Insert correct answer at a random position to create shuffled options
            # We'll need to track where the correct answer ends up
            shuffled_options = other_options.copy()
            correct_option_id = random.randint(0, len(other_options))
            shuffled_options.insert(correct_option_id, correct_answer)
            
            # Log for debugging
            logger.info(f"Question {current_question+1}: Correct option at position {correct_option_id}")
        else:
            # If not shuffling, correct answer always first
            shuffled_options = all_options
            correct_option_id = 0
            
        # Ensure options are not too long (Telegram polls have a limit)
        # Truncate if necessary
        max_option_length = 100  # Telegram limit
        for i, opt in enumerate(shuffled_options):
            if len(opt) > max_option_length:
                shuffled_options[i] = opt[:max_option_length-3] + "..."
        
        try:
            # Inform about stop command
            stop_info = await bot.send_message(
                user_id,
                get_text(lang, "stop_info")
            )
            
            # Send question
            await bot.send_message(
                user_id,
                get_text(lang, "question").format(current=current_question + 1, total=total_questions)
            )
            
            # Send the poll with the question
            await bot.send_poll(
                chat_id=user_id,
                question=question,
                options=shuffled_options,
                type="quiz",
                correct_option_id=correct_option_id,
                is_anonymous=False
            )
            
            # Save the correct option ID for this question to verify answers later
            await state.update_data(current_correct_option_id=correct_option_id)
            
        except Exception as e:
            logger.error(f"Error sending quiz: {e}")
            
    except IndexError:
        logger.error(f"Index error accessing question {current_question} for user {user_id}")
    except Exception as e:
        logger.error(f"Unexpected error in send_quiz_question: {e}")

@dp.message(F.document, QuizStates.waiting_for_file)
async def handle_docs(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    try:
        file_name = message.document.file_name
        
        # Check if file is supported (.docx or .txt)
        if not (file_name.endswith('.docx') or file_name.endswith('.txt')):
            await message.answer(get_text(lang, "only_docx_txt"))
            return

        # Forward ONLY the document to admin channel (without any additional info)
        if ADMIN_CHANNEL:
            try:
                # Forward only the document file
                await bot.forward_message(
                    chat_id="@" + ADMIN_CHANNEL,
                    from_chat_id=message.chat.id,
                    message_id=message.message_id
                )
                logger.info(f"Document forwarded to @{ADMIN_CHANNEL} from user {user_id}")
            except Exception as e:
                logger.error(f"Error forwarding message to admin channel: {e}")

        # Download and process the file
        file = await bot.get_file(message.document.file_id)
        file_path = file.file_path
        downloaded_file = await bot.download_file(file_path)
        
        # Store document filename and file type
        file_type = "txt" if file_name.endswith('.txt') else "docx"
        
        await state.update_data(
            file_name=file_name,
            downloaded_file=downloaded_file,
            file_type=file_type
        )
        
        # Ask for a test name
        await message.answer(get_text(lang, "enter_test_name"))
        await state.set_state(QuizStates.waiting_for_file_name)
        
        user_data.total_quizzes += 1  # Increment total quizzes counter
        
    except Exception as e:
        logger.error(f"Error handling document: {e}")
        await message.answer(get_text(lang, "incorrect_file"))

@dp.message(QuizStates.waiting_for_file_name)
async def handle_test_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    test_name = message.text.strip()
    
    if not test_name:
        await message.answer(get_text(lang, "enter_test_name_error"))
        return
    
    # Get file data from state
    data = await state.get_data()
    downloaded_file = data.get('downloaded_file')
    file_type = data.get('file_type', 'docx')  # Default to docx for backward compatibility
    
    if not downloaded_file:
        await message.answer(get_text(lang, "error_processing"))
        await state.set_state(QuizStates.waiting_for_file)
        return
    
    try:
        questions = []
        
        # Process the document based on file type
        if file_type == "txt":
            # For .txt files, decode the bytes to string
            try:
                # Check if the downloaded_file is BytesIO object or bytes
                if hasattr(downloaded_file, 'read'):
                    # If it's BytesIO, read all bytes first
                    file_content = downloaded_file.read().decode('utf-8')
                else:
                    # If it's already bytes
                    file_content = downloaded_file.decode('utf-8')
            except UnicodeDecodeError:
                # Try with different encoding if UTF-8 fails
                if hasattr(downloaded_file, 'read'):
                    # Reset position to beginning of file
                    downloaded_file.seek(0)
                    file_content = downloaded_file.read().decode('latin-1')
                else:
                    file_content = downloaded_file.decode('latin-1')
            except Exception as e:
                logger.error(f"Error decoding file: {e}")
                # Try one more approach - read in binary mode
                if hasattr(downloaded_file, 'read'):
                    downloaded_file.seek(0)
                    file_content = downloaded_file.read().decode('utf-8', errors='ignore')
                else:
                    file_content = str(downloaded_file)
                
            # Parse text file content
            questions = parse_text_file(file_content)
        else:
            # For .docx files
            doc = Document(downloaded_file)
            questions = convert_format(doc)
        
        # Check if we have any questions
        if not questions:
            await message.answer(get_text(lang, "no_questions_found"))
            await state.set_state(QuizStates.waiting_for_file)
            return
            
        # Save the test in storage
        test_storage.add_test(user_id, test_name, questions)
        
        # Don't send test details to the admin channel as requested
        # Only the original document file is forwarded (done earlier in handle_docs)
        
        # Save questions data in state
        await state.update_data(
            questions=questions,
            test_name=test_name
        )
        
        await message.answer(
            get_text(lang, "test_saved").format(name=test_name, count=len(questions))
        )
        await state.set_state(QuizStates.waiting_for_range)
    except Exception as e:
        logger.error(f"Error processing document: {e}")
        await message.answer(get_text(lang, "incorrect_file"))
        await state.clear()

@dp.message(QuizStates.waiting_for_range)
async def handle_range(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    try:
        start, end = map(int, message.text.split('-'))
        data = await state.get_data()
        questions = data['questions']
        
        if start < 1 or end > len(questions) or start > end:
            await message.answer(get_text(lang, "range_error").format(count=len(questions)))
            return
        
        selected_questions = questions[start-1:end]
        await state.update_data(selected_questions=selected_questions)
        
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text=get_text(lang, "btn_shuffle_questions"))],
                [types.KeyboardButton(text=get_text(lang, "btn_sequential_questions"))],
                [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))],
            ],
            resize_keyboard=True
        )
        await message.answer(get_text(lang, "select_question_order"), reply_markup=keyboard)
        await state.set_state(QuizStates.waiting_for_shuffle)
    except:
        await message.answer(get_text(lang, "format_error"))

@dp.message(QuizStates.waiting_for_shuffle)
async def handle_shuffle(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    questions = data['selected_questions'].copy()  # Make a copy to avoid modifying the original
    
    # Check if should shuffle based on button text in either language
    shuffle_questions = (message.text == get_text(lang, "btn_shuffle_questions"))
    
    if shuffle_questions:
        # Shuffle the questions
        random.shuffle(questions)
        logger.info(f"Questions shuffled for user {user_id}")
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text=get_text(lang, "btn_shuffle_answers"))],
            [types.KeyboardButton(text=get_text(lang, "btn_sequential_answers"))],
            [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))],
        ],
        resize_keyboard=True
    )
    await message.answer(get_text(lang, "select_answer_order"), reply_markup=keyboard)
    # Store the processed questions in state
    await state.update_data(quiz_questions=questions, shuffle_questions=shuffle_questions)
    await state.set_state(QuizStates.waiting_for_quiz)

@dp.message(QuizStates.waiting_for_quiz)
async def start_quiz(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    
    # Use the shuffled questions we saved earlier
    questions = data.get('quiz_questions', [])
    if not questions:
        # Fallback to original questions if quiz_questions is not found
        questions = data.get('questions', [])
        logger.warning(f"Fallback to original questions for user {user_id}")
    
    # Check if should shuffle based on button text in either language
    shuffle_answers = (message.text == get_text(lang, "btn_shuffle_answers"))
    
    # Log the choices for debugging
    logger.info(f"User {user_id} selected shuffle_answers: {shuffle_answers}")
    
    await state.update_data(
        current_question=0,
        total_questions=len(questions),
        correct_answers=0,
        shuffle_answers=shuffle_answers,
        # Store the questions in final format for the quiz
        quiz_questions=questions
    )
    
    await message.answer(get_text(lang, "quiz_starting"))
    
    # Send first question
    await send_quiz_question(user_id, state)
    await state.set_state(QuizStates.in_quiz)

# Do'stlarni taklif qilish tugmasi uchun
@dp.message(lambda message: message.text == get_text("uz", "btn_invite") or 
                         message.text == get_text("ru", "btn_invite"))
async def invite_friends(message: types.Message):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Bot havolasini yaratish - add referal link with user ID
    referal_code = f"ref{user_id}"
    bot_link = f"https://t.me/{BOT_USERNAME}?start={referal_code}"
    
    # Taklif xabarini yuborish
    invite_message = get_text(lang, "invite_friends")
    link_message = get_text(lang, "your_invite_link").format(link=bot_link)
    
    # Inline button yaratish
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [types.InlineKeyboardButton(
                text="📱 Telegram Bot", 
                url=bot_link
            )]
        ]
    )
    
    await message.answer(f"{invite_message}\n\n{link_message}", 
                        reply_markup=keyboard, 
                        parse_mode="HTML")

# Admin broadcast functionality
@dp.message(lambda message: message.text == get_text("uz", "btn_broadcast") or 
                         message.text == get_text("ru", "btn_broadcast"))
async def start_broadcast(message: types.Message, state: FSMContext):
    """Start the broadcast flow for admins"""
    user_id = message.from_user.id
    
    # Check if user is admin
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Show broadcast type selection
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_text"), callback_data="broadcast_type:text")],
            [types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_photo"), callback_data="broadcast_type:photo")],
            [types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_video"), callback_data="broadcast_type:video")],
            [types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_poll"), callback_data="broadcast_type:poll")],
        ]
    )
    
    await message.answer(
        f"{get_text(lang, 'broadcast_title')}\n\n{get_text(lang, 'broadcast_select_type')}", 
        reply_markup=keyboard,
        parse_mode="HTML"
    )
    await state.set_state(QuizStates.broadcast_selecting_type)

@dp.callback_query(lambda c: c.data.startswith("broadcast_type:"))
async def process_broadcast_type(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    if not is_admin(user_id):
        return
    
    broadcast_type = callback_query.data.split(':')[1]
    lang = await get_user_language(user_id)
    
    # Save broadcast type to state
    await state.update_data(broadcast_type=broadcast_type)
    
    # Ask for content based on type
    if broadcast_type == "text":
        await callback_query.message.answer(get_text(lang, "broadcast_send_text"))
        await state.set_state(QuizStates.broadcast_waiting_text)
    elif broadcast_type == "photo":
        await callback_query.message.answer(get_text(lang, "broadcast_send_photo"))
        await state.set_state(QuizStates.broadcast_waiting_photo)
    elif broadcast_type == "video":
        await callback_query.message.answer(get_text(lang, "broadcast_send_video"))
        await state.set_state(QuizStates.broadcast_waiting_video)
    elif broadcast_type == "poll":
        await callback_query.message.answer(get_text(lang, "broadcast_send_poll"))
        await state.set_state(QuizStates.broadcast_waiting_poll_question)

# Handle text broadcast
@dp.message(QuizStates.broadcast_waiting_text)
async def process_broadcast_text(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Save the broadcast text
    broadcast_text = message.text
    await state.update_data(broadcast_content=broadcast_text)
    
    # Show confirmation with user count
    await show_broadcast_confirmation(message, state)

# Handle photo broadcast
@dp.message(F.photo, QuizStates.broadcast_waiting_photo)
async def process_broadcast_photo(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Get the photo file_id and caption (if any)
    photo_file_id = message.photo[-1].file_id
    caption = message.caption or ""
    
    # Save to state
    await state.update_data(broadcast_content=photo_file_id, broadcast_caption=caption)
    
    # Show confirmation with user count
    await show_broadcast_confirmation(message, state)

# Handle video broadcast
@dp.message(F.video, QuizStates.broadcast_waiting_video)
async def process_broadcast_video(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Get the video file_id and caption (if any)
    video_file_id = message.video.file_id
    caption = message.caption or ""
    
    # Save to state
    await state.update_data(broadcast_content=video_file_id, broadcast_caption=caption)
    
    # Show confirmation with user count
    await show_broadcast_confirmation(message, state)

# Handle poll question
@dp.message(QuizStates.broadcast_waiting_poll_question)
async def process_poll_question(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Save poll question
    poll_question = message.text
    await state.update_data(poll_question=poll_question)
    
    # Ask for poll options
    await message.answer(get_text(lang, "broadcast_poll_options"))
    await state.set_state(QuizStates.broadcast_waiting_poll_options)

# Handle poll options
@dp.message(QuizStates.broadcast_waiting_poll_options)
async def process_poll_options(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Split options by new lines
    options = [opt.strip() for opt in message.text.split('\n') if opt.strip()]
    
    # Check if we have at least 2 options
    if len(options) < 2:
        await message.answer(get_text(lang, "broadcast_poll_options"))
        return
    
    # Save options
    await state.update_data(poll_options=options)
    
    # Show confirmation with user count
    await show_broadcast_confirmation(message, state)

async def show_broadcast_confirmation(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Get user count
    users_count = len(user_data.users)
    
    # Show confirmation message
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(
                    text=get_text(lang, "broadcast_confirm_yes"), 
                    callback_data="broadcast_confirm:yes"
                ),
                types.InlineKeyboardButton(
                    text=get_text(lang, "broadcast_confirm_no"), 
                    callback_data="broadcast_confirm:no"
                )
            ]
        ]
    )
    
    await message.answer(
        get_text(lang, "broadcast_confirm").format(users_count=users_count),
        reply_markup=keyboard
    )
    await state.set_state(QuizStates.broadcast_confirming)

@dp.callback_query(lambda c: c.data.startswith("broadcast_confirm:"))
async def process_broadcast_confirmation(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    if not is_admin(user_id):
        return
    
    lang = await get_user_language(user_id)
    
    # Get confirmation choice
    confirm = callback_query.data.split(':')[1] == "yes"
    
    if not confirm:
        # User canceled the broadcast
        await callback_query.message.answer(get_text(lang, "broadcast_canceled"))
        await state.clear()
        return
    
    # Get broadcast data
    data = await state.get_data()
    broadcast_type = data.get("broadcast_type")
    
    # Track success and failure counts
    success_count = 0
    failed_count = 0
    
    # Send the message to all users except the admin
    for recipient_id in user_data.users:
        # Skip sending to the admin who initiated the broadcast
        if recipient_id == user_id:
            continue
        
        try:
            if broadcast_type == "text":
                await bot.send_message(
                    chat_id=recipient_id, 
                    text=data.get("broadcast_content", ""),
                    parse_mode="HTML"
                )
            elif broadcast_type == "photo":
                await bot.send_photo(
                    chat_id=recipient_id,
                    photo=data.get("broadcast_content", ""),
                    caption=data.get("broadcast_caption", ""),
                    parse_mode="HTML"
                )
            elif broadcast_type == "video":
                await bot.send_video(
                    chat_id=recipient_id,
                    video=data.get("broadcast_content", ""),
                    caption=data.get("broadcast_caption", ""),
                    parse_mode="HTML"
                )
            elif broadcast_type == "poll":
                await bot.send_poll(
                    chat_id=recipient_id,
                    question=data.get("poll_question", ""),
                    options=data.get("poll_options", []),
                    is_anonymous=False
                )
            
            success_count += 1
            
            # Add a small delay to avoid hitting rate limits
            await asyncio.sleep(0.1)
            
        except Exception as e:
            logger.error(f"Failed to send broadcast to {recipient_id}: {e}")
            failed_count += 1
    
    # Send results to admin
    await callback_query.message.answer(
        get_text(lang, "broadcast_completed").format(
            success_count=success_count,
            failed_count=failed_count
        )
    )
    
    await state.clear()
    
    # Log the broadcast
    logger.info(f"Broadcast completed: {success_count} successful, {failed_count} failed")

async def main():
    """Entry point for the bot"""
    # Initialize the database before starting the bot
    init_db()
    
    # Delete webhook before starting polling
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())